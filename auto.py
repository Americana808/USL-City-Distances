from openpyxl import load_workbook
from math import radians, sin, cos, sqrt, atan2
import pytz
from datetime import datetime
import os

file_name = "/Users/Marco/Documents/.code/python/Excel auto/Marco_USL_22.xlsx"

if os.path.exists(file_name):
    workbook = load_workbook(file_name)
else:
    print(f"The file '{file_name}' does not exist in the current directory.")

sheet = workbook.active
cell_value = sheet.cell(row=1, column=1).value
print(cell_value)

cityData = {
    "ATL": {
        "lat": 33.7488,
        "lon": 84.3877,
        "timezone": "America/New_York"
    },
    "BHM": {
        "lat": 33.5186,
        "lon": 86.8104,
        "timezone": "America/Chicago"
    },
    "CHS": {
        "lat": 32.7765,
        "lon": 79.9311,
        "timezone": "America/New_York"
    },
    "COS": {
        "lat": 38.8339,
        "lon": 104.8214,
        "timezone": "America/Denver"
    },
    "DET": {
        "lat": 42.3314,
        "lon": 83.0458,
        "timezone": "America/Detroit"
    },
    "TUL": {
        "lat": 36.1540,
        "lon": 95.9928,
        "timezone": "America/Chicago"
    },
    "ELP": {
        "lat": 31.7619,
        "lon": 106.4850,
        "timezone": "America/Denver"
    },
    "HFD": {
        "lat": 41.7658,
        "lon": 72.6734,
        "timezone": "America/New_York"
    },
    "IND": {
        "lat": 39.7684,
        "lon": 86.1581,
        "timezone": "America/Indiana/Indianapolis"
    },
    "LAG": {
        "lat": 34.0549,
        "lon": 118.2426,
        "timezone": "America/Los_Angeles"
    },
    "LV": {
        "lat": 36.1716,
        "lon": 115.1391,
        "timezone": "America/Los_Angeles"
    },
    "LDN": {
        "lat": 39.1155,
        "lon": 77.5645,
        "timezone": "America/New_York"
    },
    "LOU": {
        "lat": 38.2527,
        "lon": 85.7585,
        "timezone": "America/New_York"
    },
    "MEM": {
        "lat": 35.1495,
        "lon": 90.0490,
        "timezone": "America/Chicago"
    },
    "MIA": {
        "lat": 25.7617,
        "lon": 80.1918,
        "timezone": "America/New_York"
    },
    "MB": {
        "lat": 36.8007,
        "lon": 121.9473,
        "timezone": "America/Los_Angeles"
    },
    "NM": {
        "lat": 35.0844,
        "lon": 106.6504,
        "timezone": "America/Denver"
    },
    "NYRB": {
        "lat": 40.8259,
        "lon": 74.2090,
        "timezone": "America/New_York"
    },
    "OAK": {
        "lat": 37.8044,
        "lon": 122.2712,
        "timezone": "America/Los_Angeles"
    },
    "OC": {
        "lat": 33.7175,
        "lon": 117.8311,
        "timezone": "America/Los_Angeles"
    },
    "PHX": {
        "lat": 33.4484,
        "lon": 112.0740,
        "timezone": "America/Phoenix"
    },
    "PIT": {
        "lat": 40.4406,
        "lon": 79.9959,
        "timezone": "America/New_York"
    },
    "RGV": {
        "lat": 26.22,
        "lon": 98.12,
        "timezone": "America/Chicago"
    },
    "SAC": {
        "lat": 38.5816,
        "lon": 121.4944,
        "timezone": "America/Los_Angeles"
    },
    "SA": {
        "lat": 29.4252,
        "lon": 98.4946,
        "timezone": "America/Chicago"
    },
    "SD": {
        "lat": 32.7157,
        "lon": 117.1611,
        "timezone": "America/Los_Angeles"
    },
    "TBR": {
        "lat": 27.7634,
        "lon": 82.5437,
        "timezone": "America/New_York"
    }
}

def calculateDistance(lat1, lon1, lat2, lon2):
    # Convert latitude and longitude from degrees to radians
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])

    # Radius of the Earth in kilometers
    radius = 6371
    
    # Haversine formula
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    distance = radius * c
    
    return distance

i = 0
while i < 460:
    home = sheet.cell(row=i+1, column=3).value
    away = sheet.cell(row=i+1, column=7).value

    if home in cityData and away in cityData:
        lat1 = cityData[home]["lat"]
        lon1 = cityData[home]["lon"]
        lat2 = cityData[away]["lat"]
        lon2 = cityData[away]["lon"]

        distance = calculateDistance(lat1, lon1, lat2, lon2)

        sheet.cell(row=i+1, column=17).value = distance

        """
        homeTimezone = pytz.timezone(cityData[home]["timezone"])
        awayTimezone = pytz.timezone(cityData[away]["timezone"])

        currentHomeTime = datetime.now(homeTimezone)
        currentAwayTime = datetime.now(awayTimezone)

        time_difference = currentHomeTime - currentAwayTime

        hoursDifference = time_difference.days * 24 + time_difference.seconds // 3600

        print(currentHomeTime)
        sheet.cell(row=i+1, column=23).value = f"{hoursDifference}"
        """

    i+=1

workbook.save(file_name)
# MODIFYING CELLS
# sheet.cell(row=1, column=1).value = "New Value"
